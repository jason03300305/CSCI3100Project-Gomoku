import pygame

class Button:
    def __init__(self, screen, pos, text, font, base_color, hov_color):
        self.x_pos = pos[0]
        self.y_pos = pos[1]
        self.font = font
        self.base_color, self.hov_color = base_color, hov_color
        self.text_input = text
        self.text = self.font.render(text, True, self.base_color)
        self.rect = self.text.get_rect(center=(self.x_pos, self.y_pos))
        self.text_rect = self.rect
        self.draw(screen)
    
    def draw(self, screen):
        screen.blit(self.text, self.text_rect)

    def checkMousePos(self, pos):
        # Check if the given position is within the button's bounds
        if self.rect.left <= pos[0] <= self.rect.right and self.rect.bottom >= pos[1] >= self.rect.top:
            return True
        return False
        
    def changeColor(self, pos):
        # Change the button's text color if hovering on the button
        if self.rect.left <= pos[0] <= self.rect.right and self.rect.bottom >= pos[1] >= self.rect.top:
            self.text = self.font.render(self.text_input, True, self.hov_color)
        else:
            self.text = self.font.render(self.text_input, True, self.base_color)

pygame.init()  # Initialize Pygame

screen = pygame.display.set_mode((1280, 720))  # Set up the screen
pygame.display.set_caption("Player List")  # Set the window caption

menu_font = pygame.font.Font('Cartoon.ttf', 100)
default_font = pygame.font.Font('Cartoon.ttf', 50)
default_font2 = pygame.font.Font('Cartoon.ttf', 28)

while True:
    screen.fill('black')  # Clear the screen with black color
    mouse_pos = pygame.mouse.get_pos()

    menu_text = menu_font.render("PLAYER LIST", True, "white")  # Render the menu text
    menu_rect = menu_text.get_rect(center=(640, 100))  # Get the rectangle for the menu text and center it
    screen.blit(menu_text, menu_rect)  # Draw the menu text on the screen



    import openpyxl

    # Load the workbook
    workbook = openpyxl.load_workbook('/Users/wanpuichoi/Library/Mobile Documents/com~apple~CloudDocs/Year3/Sem2/CSCI3100/project/GOMOKU/player_list.xlsx')

    # Select the active sheet
    sheet = workbook.active

    # Get the dimensions of the table
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Define the position and size of the table
    table_x = 355
    table_y = 200
    cell_width = 150
    cell_height = 30

    # Loop through the table data and draw each cell
    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell_value = sheet.cell(row=row, column=col).value
            cell_x = table_x + (col - 1) * cell_width
            cell_y = table_y + (row - 1) * cell_height
 
            cell_rect = pygame.Rect(cell_x, cell_y, cell_width, cell_height)

            pygame.draw.rect(screen, 'black', cell_rect)
            cell_text = default_font2.render(str(cell_value), True, 'white')
            cell_text_rect = cell_text.get_rect(center=cell_rect.center)
            screen.blit(cell_text, cell_text_rect)

    # Save the changes to the workbook
    workbook.save('/Users/wanpuichoi/Library/Mobile Documents/com~apple~CloudDocs/Year3/Sem2/CSCI3100/project/GOMOKU/player_list.xlsx')

    a = 400
    d = 70

    # Create buttons and draw them on the screen
    btn_edit = Button(screen, pos=(256, a+d*4), text='EDIT', font=default_font, base_color='gray', hov_color='white')
    btn_back = Button(screen, pos=(1100, a+d*4), text='BACK', font=default_font, base_color='gray', hov_color='white')


    # Change button colors based on mouse position and draw them
    for btn in [btn_edit, btn_back]:
        btn.changeColor(mouse_pos)
        btn.draw(screen)

    # Handle events
    for event in pygame.event.get():
        if event.type == pygame.QUIT:  # If the user closes the window, quit the game
            pygame.quit()
        if event.type == pygame.MOUSEBUTTONDOWN:
            # Check which button is clicked and perform corresponding actions
            if btn_edit.checkMousePos(mouse_pos):
                
                print(f"{btn_edit.text_input} is pressed.")

            if btn_back.checkMousePos(mouse_pos):
                print(f"{btn_back.text_input} is pressed.")
                pygame.quit()

   
    pygame.display.update()  # Update the display to show the changes